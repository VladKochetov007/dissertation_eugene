import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors & Styles ──
HEADER_FILL = PatternFill('solid', fgColor='1B2A4A')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='E8EDF3')
SECTION_FONT = Font(name='Arial', bold=True, color='1B2A4A', size=10)
DATA_FONT = Font(name='Arial', size=9, color='333333')
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top', horizontal='left')
CENTER_ALIGN = Alignment(wrap_text=True, vertical='top', horizontal='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# Priority color fills
P5_FILL = PatternFill('solid', fgColor='C6EFCE')  # green - highest
P4_FILL = PatternFill('solid', fgColor='DFF0D8')
P3_FILL = PatternFill('solid', fgColor='FFF2CC')  # yellow - medium
P2_FILL = PatternFill('solid', fgColor='FCE4D6')
P1_FILL = PatternFill('solid', fgColor='F4CCCC')  # red - lowest
P0_FILL = PatternFill('solid', fgColor='EEEEEE')

PRIORITY_FILLS = {5: P5_FILL, 4: P4_FILL, 3: P3_FILL, 2: P2_FILL, 1: P1_FILL, 0: P0_FILL}

# ── Column structure ──
COLUMNS = [
    ('A', 'Priority\n(0-5)', 8),
    ('B', 'Firm', 22),
    ('C', 'Vertical', 18),
    ('D', 'Why Causal ML?', 42),
    ('E', 'Target Role', 20),
    ('F', 'Hook / Stand Out', 42),
    ('G', 'Personalized Outreach Draft', 60),
    ('H', 'Status', 14),
    ('I', 'Notes', 30),
]

# ── Data ──
prospects = [
    # Section: ART & LUXURY COLLECTIBLES
    ('SECTION', 'ART & LUXURY COLLECTIBLES'),
    (5, 'Phillips', 'Art Auctions',
     'Pre-sale estimate accuracy is the core business problem. Causal ML identifies what actually drives hammer prices beyond comp-based correlations — provenance chains, exhibition history, collector migration patterns.',
     'Head of Data / Chief Commercial Officer',
     'New CEO Stephen Brooks (ex-Christie\'s CFO, appointed 2025) signals operational efficiency push. Fresh leadership = openness to new tools. Bloomsbury\'s existing art market work is directly relevant proof.',
     'Hi [Name], I noticed Phillips\' leadership transition under Stephen Brooks — exciting times. We\'re Bloomsbury Technology, a London causal AI company working with major art market players to improve valuation accuracy. Our models identify what actually drives hammer prices (not just comps) — provenance, exhibition history, collector migration. Would love 20 minutes to show you what we\'ve found. Happy to share a quick case study ahead of time.',
     '', ''),
    (4, 'Sotheby\'s', 'Art Auctions',
     'Post-Cappellazzo restructuring in fine art division. Massive transaction data but pricing still relies on specialist intuition. Causal ML could systematize what the best specialists do intuitively.',
     'VP of Analytics / Head of Fine Art',
     'Structural reorganization post-Amy Cappellazzo departure (2024) suggests openness to new analytical tools. Sotheby\'s is digitizing its sales process — causal ML fits naturally into that pipeline.',
     'Hi [Name], Sotheby\'s digitization push is impressive. We\'re working with art market leaders on something complementary — causal AI that reveals WHY certain lots outperform estimates, beyond what correlational models catch. We\'ve found some surprising drivers in contemporary and post-war segments. Would a brief walkthrough be useful?',
     '', ''),
    (3, 'Beaumont Nathan', 'Art Advisory',
     'Top-end advisory serving UHNWIs. Opaque market intelligence — need causal analysis showing why certain contemporary/Old Master markets move (estate news, museum acquisitions, wealth migration).',
     'Managing Partner / Head of Research',
     'Advises at "top end of market" with offices in London, NY, Abu Dhabi. Cross-geography presence = data advantage if paired with causal tools. Their clients demand rigorous rationale for 7-8 figure purchases.',
     'Hi [Name], advising at your level requires more than intuition — your clients expect rigorous rationale. We build causal AI that identifies what actually drives market movements in contemporary and Old Master segments: estate releases, museum deaccessions, collector migration, currency effects. Could be a powerful complement to your advisory. Worth a conversation?',
     '', ''),
    (3, 'Lyon & Turnbull', 'Art Auctions',
     'High-volume multi-category house (contemporary, design, ceramics). Batch valuation at scale = perfect for causal ML. London Mall Galleries location.',
     'Managing Director / Head of Valuations',
     'UK specialist with strong regional presence. Multi-category means causal factors differ by segment — one model doesn\'t fit all. Bloomsbury can show cross-category causal insights.',
     'Hi [Name], running multi-category auctions means pricing drivers vary enormously between contemporary, design, and decorative arts. We\'ve built causal ML tools that identify segment-specific value drivers — what actually causes price movements in each category. Could be useful for your valuation and estimate processes. Happy to show you some examples.',
     '', ''),

    # Section: WINE
    ('SECTION', 'WINE INVESTMENT & TRADING'),
    (5, 'WineFi', 'Wine Investment',
     'Already quantitative (18M+ data points, 2000 producers) but correlational only. They\'re 50% there — they need the causal layer. What CAUSES price appreciation vs. what just correlates?',
     'Chief Investment Officer / Head of Data Science',
     'They explicitly brand as "world\'s first quantitative fine wine investment firm." They WANT to be data-driven but are missing the causal dimension. This is the single most natural fit on the entire list.',
     'Hi [Name], WineFi\'s quantitative approach is exactly right — 18M data points is serious. But there\'s a meaningful gap between correlation and causation in wine markets. We build causal ML that identifies what actually drives price appreciation: vintage scarcity, weather patterns, collector migration, currency effects, vs. what just happens to move together. For a firm positioning as quantitative-first, this could be a real edge. 20 minutes?',
     '', ''),
    (4, 'Cru Wine', 'Wine Investment',
     '11.8% annualized gross returns. Need causal edge to maintain outperformance. Understanding what drives Burgundy vs. Bordeaux differently is a causal question, not a correlational one.',
     'VP of Investment Strategy / Head of Analytics',
     'Established track record means they care about sustainability of returns. Next competitive move is understanding WHY their winners win.',
     'Hi [Name], 11.8% annualized in fine wine is impressive — the question is what sustains it. We build causal AI that goes beyond "Burgundy outperformed last year" to explain WHY — is it scarcity dynamics, collector demographics, currency flows, or vintage effects? Understanding the causal mechanism means you can predict when the pattern will hold and when it\'ll break. Worth a conversation?',
     '', ''),
    (3, 'Liv-ex', 'Wine Exchange',
     'London-based fine wine exchange — the marketplace itself. Their pricing indices are correlational. Causal layer would transform the product. Strategic partnership potential.',
     'Head of Data / Chief Product Officer',
     'They ARE the infrastructure. If Bloomsbury\'s causal layer sits on top of Liv-ex data, it becomes the industry standard. Partnership > client relationship.',
     'Hi [Name], Liv-ex\'s pricing infrastructure is the backbone of fine wine markets. We\'ve been thinking about what a causal intelligence layer on top of that data could look like — not just what correlates with price movements, but what causes them. Could be interesting as a product extension or partnership. Worth exploring?',
     '', ''),

    # Section: WATCHES & CLASSIC CARS
    ('SECTION', 'WATCHES & CLASSIC CARS'),
    (4, 'Talacrest', 'Classic Ferraris',
     'World\'s #1 classic Ferrari dealer (1600+ cars sold). Massive transaction history. Causal analysis of rarity, provenance, racing history, ownership, economic cycles = inventory positioning edge.',
     'CEO / Chief Commercial Officer',
     'Dominant market position but data advantage is underutilized. They have the most Ferrari transaction data in the world. Causal ML turns that into predictive intelligence.',
     'Hi [Name], with 1600+ Ferraris traded, Talacrest has the deepest dataset in the classic Ferrari market. We build causal AI that turns transaction history into forward-looking intelligence — identifying what actually drives value (provenance, racing history, production rarity, economic cycles) vs. what just correlates. For inventory positioning and timing, this could be significant. Interested in a walkthrough?',
     '', ''),
    (3, 'Chrono24', 'Watch Marketplace',
     'Global watch marketplace. Pricing transparency is their value prop but it\'s correlational. Causal layer = product differentiator. What CAUSES watch price movements?',
     'Head of Data Science / VP Product',
     'Platform play — similar logic to Liv-ex. If causal intelligence sits on Chrono24\'s data, it becomes the industry pricing standard. Also: watch market had a significant correction 2023-24, making causal understanding of cycles more valuable.',
     'Hi [Name], Chrono24\'s pricing data is unmatched in watches. The opportunity: moving from "what are similar watches selling for" to "what\'s causing this reference to appreciate or depreciate." Post-2023 market correction, understanding the causal drivers of watch cycles is more valuable than ever. We build exactly this kind of causal AI. Worth exploring as a product enhancement?',
     '', ''),
    (3, 'Hexagon Classics', 'Classic Cars',
     'Largest UK/European multi-marque collection (40+ Porsches in stock alone). Scale operation where causal pricing model = margin improvement across inventory.',
     'Managing Director / Head of Operations',
     'London/Highgate location. Scale inventory means small pricing improvements compound. Causal model linking condition, mileage, model cycles, collector demographics = inventory optimization.',
     'Hi [Name], managing 40+ Porsches at one location (plus the rest of the collection) means pricing accuracy compounds across every unit. We build causal AI that identifies what actually drives value for each model — condition factors, market cycles, collector demographics, competitor pricing — beyond simple comps. Could meaningfully improve margins. Want to see how it works?',
     '', ''),

    # Section: INSURANCE
    ('SECTION', 'INSURANCE & REINSURANCE'),
    (5, 'Chaucer Group', 'Lloyd\'s Syndicate',
     'Specializes in complex, high-severity/low-frequency risks. 17+ years of policy/claims data. Already investing in data infrastructure (Trillium partnership). Next step: causal analysis of claims drivers.',
     'Chief Data Officer / Head of Underwriting Analytics',
     'Already spending on data governance (GRC software with Trillium). They\'re ready for the next layer. The pitch writes itself: "You\'ve cleaned your data. Now let\'s find what actually causes claims."',
     'Hi [Name], I saw Chaucer\'s partnership with Trillium on data governance — smart foundation. The natural next step: once your data is clean, causal ML can identify what actually CAUSES claims vs. what just correlates with them. We build exactly this — causal AI that separates genuine risk drivers from noise. For specialty lines, the impact on loss ratios can be significant. Would love to show you some examples.',
     '', ''),
    (4, 'Beazley', 'Specialty Insurance',
     'Leading specialty insurer (cyber, professional liability, marine). Cyber insurance especially needs causal understanding — what causes breaches vs. what correlates with them.',
     'Chief Risk Officer / Head of Data Science',
     'Beazley\'s cyber practice is a standout. Cyber risk pricing is notoriously difficult because correlational models fail — attack vectors change constantly. Causal understanding of breach drivers is the holy grail.',
     'Hi [Name], pricing cyber risk is one of the hardest problems in insurance — attack vectors evolve faster than correlational models can track. We build causal AI that identifies what structural factors actually cause breaches: network architecture, employee behavior, patch cadence, vendor concentration. For Beazley\'s cyber practice specifically, this could transform underwriting accuracy. Worth 20 minutes?',
     '', ''),
    (4, 'Hiscox', 'Specialty Insurance',
     'London HQ, known for innovation. Art & private client insurance is directly adjacent to Bloomsbury\'s art market work. Causal ML for art insurance pricing = natural extension.',
     'Head of Data & Analytics / Chief Underwriter (Art & Private Client)',
     'Hiscox insures high-value art collections. Bloomsbury already works in art valuation. Art insurance pricing using causal ML = perfect crossover pitch.',
     'Hi [Name], Hiscox\'s art and private client insurance sits at an interesting intersection with our work. We build causal AI for opaque asset markets — art, collectibles, luxury goods — identifying what actually drives value and risk. For art insurance specifically, causal understanding of damage risk, market volatility, and collection concentration could improve pricing significantly. Given our existing art market work, there\'s real synergy here.',
     '', ''),

    # Section: PRIVATE EQUITY
    ('SECTION', 'PRIVATE EQUITY & VENTURE CAPITAL'),
    (4, 'Graphite Capital', 'Mid-Market PE',
     'UK specialist (£10m-£100m deals). 35-person team. Need better due diligence analytics. Causal ML identifies what operational factors actually CAUSE value creation in portfolio companies.',
     'Partner / Head of Investment Operations',
     'Pure UK mid-market focus = concentrated expertise. Causal analysis of their portfolio\'s value-creation levers would directly improve returns. Small team = faster decision-making on new tools.',
     'Hi [Name], in mid-market PE, the quality of due diligence is the edge. We build causal AI that identifies what actually drives EBITDA improvement in portfolio companies — separating genuine operational levers from market tailwinds. For a focused UK investor like Graphite, this could meaningfully improve both deal selection and post-acquisition strategy. Happy to show you how it works with a relevant example.',
     '', ''),
    (3, 'Inflexion', 'Mid-Market PE',
     'European scope (London, Manchester, Amsterdam, Stockholm). Multi-geography portfolio creates data richness for cross-portfolio causal analysis.',
     'Partner / Head of Analytics',
     'Pan-European exposure means they can test whether value-creation playbooks transfer across geographies — a causal question. Scale portfolio = more data for the models.',
     'Hi [Name], with portfolio companies across the UK and Europe, Inflexion has unique data on what value-creation approaches work where. We build causal AI that identifies which operational playbooks actually CAUSE improvement vs. which just happen alongside market tailwinds. Especially useful for pattern-matching across geographies. Worth a conversation?',
     '', ''),
    (3, 'Livingbridge', 'Mid-Market PE',
     'Growing firm with mid-market portfolio. Intense competition for deals means differentiation through better analytics = sustainable edge.',
     'Partner / Chief Operating Officer',
     'Competition for mid-market deals is fierce. Causal ML as a due diligence differentiator = "we see things other PE firms can\'t." Also useful for portfolio monitoring.',
     'Hi [Name], mid-market deal competition keeps intensifying. One differentiator that\'s hard to replicate: causal AI for due diligence and portfolio monitoring. Instead of correlational pattern-matching, understanding what actually causes value creation in target companies. We\'ve built this for other opaque markets and the PE application is natural. Interested in exploring?',
     '', ''),

    # Section: SANCTIONS & COMPLIANCE
    ('SECTION', 'SANCTIONS & COMPLIANCE'),
    (4, 'OFSI-regulated commodity traders', 'Trade Compliance',
     '37 live OFSI investigations into Russian oil price cap evasion. UK consolidating sanctions list (Jan 2026). Commodity traders need causal understanding of evasion patterns vs. legitimate trade.',
     'Chief Compliance Officer / Head of Surveillance',
     'Timing is perfect: Jan 2026 UK sanctions list consolidation + active OFSI enforcement. Traders face existential compliance risk. Causal ML differentiates evasion patterns from innocent ones — reduces false positives AND catches real violations.',
     'Hi [Name], with 37 active OFSI investigations and the UK sanctions list consolidation in January, compliance pressure on commodity traders is at an all-time high. Standard screening tools generate enormous false positives. We build causal AI that identifies what trade structures actually indicate evasion vs. what just looks suspicious. The difference matters — both for catching real violations and reducing compliance costs. Worth a discussion?',
     '', ''),
    (3, 'Standard Chartered', 'AML/Sanctions',
     'Heavy Asia/Middle East/Africa exposure = complex sanctions landscape. History of compliance issues = strong motivation to invest in better tools.',
     'Head of Financial Crime / Chief Compliance Officer',
     'StanChart has paid billions in compliance fines historically. Strong institutional motivation to get ahead of the problem. New Russian energy sanctions (Oct 2025) add urgency.',
     'Hi [Name], Standard Chartered\'s global exposure — particularly in corridors affected by Russia/Iran sanctions — creates a uniquely complex compliance challenge. Correlational screening generates too many false positives in your markets. We build causal AI that distinguishes genuine sanctions evasion structures from legitimate trade patterns. Given the regulatory environment, this could be worth exploring.',
     '', ''),

    # Section: QUANT TRADING
    ('SECTION', 'QUANTITATIVE TRADING'),
    (4, 'XTX Markets', 'Algorithmic Trading',
     'London-based, 50,000+ instruments, already heavy ML users. Causal inference is the next frontier — why prices move, not just that they move. Outperforms correlational ML in structural breaks.',
     'Head of Research / Chief Data Officer',
     'XTX is one of the most sophisticated ML-driven trading firms. They\'d understand the causal inference value prop immediately. The pitch is technical: causal models are robust to distribution shifts that break correlational models.',
     'Hi [Name], XTX\'s ML-driven approach is best-in-class for correlational pattern recognition. The frontier we\'re working on: causal inference for markets — models that understand WHY prices move, not just that they do. Key advantage: robustness to regime changes and structural breaks where correlational models fail. We\'d love to share some of our research. Technical conversation welcome.',
     '', ''),
    (3, 'Flow Traders', 'Algorithmic Trading',
     'Global liquidity provision across equities, forex, fixed income. Causal understanding of order flow and spread dynamics = better inventory management.',
     'VP of Analytics / Head of Research',
     'Market-making firm where understanding what CAUSES order flow matters directly for P&L. London/Finsbury Square office. Causal ML for microstructure analysis.',
     'Hi [Name], for a market-maker, understanding what causes order flow — not just predicting it — is the edge. We build causal AI that identifies the structural drivers of spread dynamics, inventory pressure, and flow patterns. Could be interesting for Flow Traders\' multi-asset operation. Worth a technical conversation?',
     '', ''),

    # Section: PROPTECH
    ('SECTION', 'PROPTECH & REAL ESTATE'),
    (3, 'Concrete VC', 'PropTech VC',
     'Invests in early-stage PropTech (€100k-€1.5m) with strategic partners JLL and Starwood Capital. Need causal understanding of what makes PropTech companies succeed.',
     'Managing Partner',
     'VC with corporate PropTech partners. They need to understand causal drivers of PropTech success. Also: potential intro to their portfolio companies as downstream clients.',
     'Hi [Name], Concrete VC\'s thesis on PropTech is interesting — and your portfolio companies face a common challenge: separating what actually drives property value from what just correlates with it. We build causal AI for opaque asset markets. Could be useful both for your due diligence and as a capability your portfolio companies deploy. Worth exploring the intersection?',
     '', ''),
    (2, 'Pi Labs', 'PropTech VC',
     'London PropTech investor. Similar to Concrete VC — partnership potential for portfolio-wide deployment of causal ML tools.',
     'General Partner',
     'Prominent in London PropTech ecosystem. Gateway to multiple PropTech companies.',
     'Hi [Name], Pi Labs\' PropTech portfolio likely encounters a recurring challenge: property valuation models that are correlational, not causal. We build AI that identifies what actually drives property values — infrastructure, demographics, zoning, not just comps. Could be interesting both for due diligence and as a capability for portfolio companies. Coffee?',
     '', ''),

    # Section: KSE / EXISTING PIPELINE EXPANSION
    ('SECTION', 'EXISTING PIPELINE EXPANSION'),
    (5, 'Gagosian', 'Art Gallery',
     'World\'s largest gallery. If Zwirner conversation progresses, Gagosian is the natural next target. Massive artist roster, global footprint, data-rich operation.',
     'Head of Operations / Director of Technology',
     'Reference Bloomsbury\'s art market work (without naming Zwirner). Gagosian\'s scale means even small pricing improvements are worth millions. They\'re known to invest in technology.',
     'Hi [Name], we\'re a London causal AI company working with leading art market institutions. Our tools identify what actually drives sales outcomes — beyond comps and specialist intuition — using causal machine learning. At Gagosian\'s scale, even marginal pricing improvements are significant. Would love to show you what we\'ve found in the contemporary market specifically.',
     '', ''),
    (4, 'Hauser & Wirth', 'Art Gallery',
     'Major gallery with strong institutional relationships and growing data operation. Known for long-term artist development — causal understanding of what builds artist markets over decades.',
     'Director of Sales / Head of Technology',
     'H&W\'s long-term artist development approach means they care about what CAUSES sustainable market growth for artists, not just short-term auction spikes. Causal ML aligns with their philosophy.',
     'Hi [Name], Hauser & Wirth\'s approach to long-term artist market development is distinctive. We build causal AI that identifies what actually drives sustainable market growth — museum placements, critical attention, collector base diversification — vs. what just correlates with temporary price spikes. Could be a powerful complement to your market intelligence. Worth a conversation?',
     '', ''),
    (4, 'Hagerty', 'Classic Car Insurance + Data',
     'The data company behind classic car insurance and valuation. They already have a valuation tool — causal ML would be a major upgrade to their pricing engine.',
     'Head of Data Science / Chief Product Officer',
     'Hagerty is to classic cars what Liv-ex is to wine — the data infrastructure. Partnership potential, not just client relationship. Their valuation tool could be transformed by causal layer.',
     'Hi [Name], Hagerty\'s valuation data is the industry standard for classic cars. We build causal AI that could take that a step further — identifying what actually drives value changes (not just comps) across marques, conditions, and market cycles. Could be interesting as a product enhancement or research partnership. Worth exploring?',
     '', ''),
]

# ── Build the sheet ──
ws = wb.active
ws.title = 'Prospects'

# Set column widths
for col_letter, header, width in COLUMNS:
    ws.column_dimensions[col_letter].width = width

# Header row
for i, (col_letter, header, _) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=i, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'
ws.auto_filter.ref = 'A1:I1'

row_num = 2
for item in prospects:
    if item[0] == 'SECTION':
        # Section header row
        ws.merge_cells(f'A{row_num}:I{row_num}')
        cell = ws.cell(row=row_num, column=1, value=item[1])
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(vertical='center', horizontal='left')
        cell.border = THIN_BORDER
        for c in range(2, 10):
            ws.cell(row=row_num, column=c).fill = SECTION_FILL
            ws.cell(row=row_num, column=c).border = THIN_BORDER
        ws.row_dimensions[row_num].height = 24
        row_num += 1
        continue

    priority, firm, vertical, why, role, hook, draft, status, notes = item
    values = [priority, firm, vertical, why, role, hook, draft, status, notes]

    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.font = DATA_FONT
        cell.alignment = WRAP_ALIGN if i >= 4 else Alignment(wrap_text=True, vertical='top')
        cell.border = THIN_BORDER

    # Priority cell formatting
    p_cell = ws.cell(row=row_num, column=1)
    p_cell.alignment = CENTER_ALIGN
    if priority in PRIORITY_FILLS:
        p_cell.fill = PRIORITY_FILLS[priority]
    p_cell.font = Font(name='Arial', bold=True, size=11, color='333333')

    # Status column - default to "Not started"
    s_cell = ws.cell(row=row_num, column=8)
    if not s_cell.value:
        s_cell.value = 'Not started'
    s_cell.alignment = CENTER_ALIGN

    ws.row_dimensions[row_num].height = 100
    row_num += 1

# ── Summary sheet ──
summary = wb.create_sheet('Summary')
summary.column_dimensions['A'].width = 25
summary.column_dimensions['B'].width = 12
summary.column_dimensions['C'].width = 45

summary_headers = ['Vertical', 'Count', 'Top Priority Target']
for i, h in enumerate(summary_headers, 1):
    cell = summary.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

verticals = [
    ('Art & Luxury Collectibles', 4, 'Phillips (new CEO, operational efficiency push)'),
    ('Wine Investment & Trading', 3, 'WineFi (already quantitative, needs causal layer)'),
    ('Watches & Classic Cars', 3, 'Talacrest (1600+ Ferraris, deepest dataset)'),
    ('Insurance & Reinsurance', 3, 'Chaucer Group (already investing in data infra)'),
    ('Private Equity', 3, 'Graphite Capital (UK mid-market, small team)'),
    ('Sanctions & Compliance', 2, 'OFSI-regulated traders (Jan 2026 deadline)'),
    ('Quantitative Trading', 2, 'XTX Markets (sophisticated, understands ML)'),
    ('PropTech & Real Estate', 2, 'Concrete VC (gateway to portfolio companies)'),
    ('Existing Pipeline Expansion', 3, 'Gagosian (natural next after Zwirner)'),
]

for i, (vert, count, top) in enumerate(verticals, 2):
    summary.cell(row=i, column=1, value=vert).font = DATA_FONT
    summary.cell(row=i, column=2, value=count).font = DATA_FONT
    summary.cell(row=i, column=3, value=top).font = DATA_FONT
    for c in range(1, 4):
        summary.cell(row=i, column=c).alignment = WRAP_ALIGN
        summary.cell(row=i, column=c).border = THIN_BORDER

total_row = len(verticals) + 2
summary.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
summary.cell(row=total_row, column=2, value=f'=SUM(B2:B{total_row-1})')
summary.cell(row=total_row, column=2).font = Font(name='Arial', bold=True, size=10)
for c in range(1, 4):
    summary.cell(row=total_row, column=c).border = THIN_BORDER

# ── Priority Guide sheet ──
guide = wb.create_sheet('Priority Guide')
guide.column_dimensions['A'].width = 10
guide.column_dimensions['B'].width = 60

guide_data = [
    ('Priority', 'Meaning'),
    (5, 'Immediate outreach — strong fit, warm angle, or time-sensitive opportunity'),
    (4, 'High priority — strong fit, clear value prop, reach out within 2 weeks'),
    (3, 'Medium priority — good fit, needs more research or a warm intro'),
    (2, 'Lower priority — interesting but longer sales cycle or less clear fit'),
    (1, 'Speculative — worth tracking but not active outreach yet'),
    (0, 'Parked — revisit later or deprioritize'),
]

for i, (p, meaning) in enumerate(guide_data):
    if i == 0:
        for j, val in enumerate([p, meaning], 1):
            cell = guide.cell(row=1, column=j, value=val)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
    else:
        cell_a = guide.cell(row=i+1, column=1, value=p)
        cell_a.font = Font(name='Arial', bold=True, size=11)
        cell_a.alignment = CENTER_ALIGN
        cell_a.border = THIN_BORDER
        if p in PRIORITY_FILLS:
            cell_a.fill = PRIORITY_FILLS[p]

        cell_b = guide.cell(row=i+1, column=2, value=meaning)
        cell_b.font = DATA_FONT
        cell_b.alignment = WRAP_ALIGN
        cell_b.border = THIN_BORDER

OUT = '/sessions/zen-clever-babbage/mnt/maynard/bloomsbury_gtm_prospects.xlsx'
wb.save(OUT)
print(f'Saved to {OUT}')
